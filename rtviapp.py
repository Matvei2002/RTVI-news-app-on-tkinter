import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter.ttk import *
from tkinter.scrolledtext import *
from PIL import Image, ImageTk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLUE


class NewsApp:
    def __init__(self, root):
        self.root = root
        self.root.geometry("800x480")
        self.root.resizable(0, 0)
        self.root.config(bg='#FF9999')
        self.root.title("Новости RTVI")
        icon = tk.PhotoImage(file='images/rtvi.png')
        self.root.iconphoto(False, icon)

        self.setup_ui()
        self.news = []

    def setup_ui(self):
        self.create_logo_frame()
        self.create_title_frame()
        self.create_input_widgets()
        self.create_parse_button()
        self.create_news_display()

    def create_logo_frame(self):
        logo_image = Image.open("images/rtvi1.png")
        logo_image = ImageTk.PhotoImage(logo_image)
        logo_label = tk.Label(self.root, image=logo_image, background = '#FF9999')
        logo_label.image = logo_image  
        logo_label.pack()

    def create_title_frame(self):
        label_frame = tk.Frame(self.root)
        label_frame.pack()
        label = tk.Label(label_frame, text="Новости", font="Arial 20 bold", bg='#FF9999', width=500)
        label.pack(side='left')

    def create_input_widgets(self):
        self.entry = tk.Entry(self.root, font=('Arial 10'))
        self.entry.pack(side="top")
        self.entry.insert(0, 'https://rtvi.com/news/')

        input_label = tk.Label(self.root, text="Количество новостей (от 1 до 10):", font="Arial 10 bold", bg='#FF9999')
        input_label.pack()
        self.input_entry = tk.Entry(self.root, width=10, font=('Arial 10'))
        self.input_entry.pack(side="top")

    def create_news_display(self):
        text_frame = tk.Frame(self.root)
        text_frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(text_frame)
        canvas.pack(side="left", fill="both", expand=True)

        self.text = tk.Text(canvas, font=('Times 14'))
        self.text.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(text_frame, command=self.text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.text.config(yscrollcommand=scrollbar.set)
        
    def create_parse_button(self):
        style = Style()
        style.configure('TButton', font=('Arial 10 bold'), foreground='#DC143C', background='#FF9999')
        button = Button(self.root, text="Поиск новостей", style='TButton', command=self.on_button_click)
        button.pack(side="top", pady=10, padx=20)

    def get_news(self, url, num=10):
        response = requests.get(url)
        html_doc = response.text
        soup = BeautifulSoup(html_doc, "html.parser")

        news = []
        
        for news_elem in soup.find_all('div', class_='arch-block')[:num]:
            title_elem = news_elem.find('h2', class_='arch-title')
            title = title_elem.text.strip()

            datetime_elem = news_elem.find('div', class_='date')
            datetime = datetime_elem.text.strip()

            link_elem = news_elem.find('a', href=True)
            link = link_elem['href']

            news.append({'title': title, 'datetime': datetime, 'link': link})

        return news

    def on_button_click(self):
        url = self.entry.get()
        num_input = self.input_entry.get()

        if not num_input.isdigit():  # Проверяем, является ли введенное значение числом
            self.text.delete('1.0', tk.END)
            self.text.insert(tk.END, "Ошибка: Введите целое положительное число в поле количества новостей")
            return
        
        num = int(num_input)
        if num < 1 or num > 10:
            self.text.delete('1.0', tk.END)
            self.text.insert(tk.END, "Ошибка: Введите число от 1 до 10")
            return
        else:
            self.news = self.get_news(url, num=num)
            self.display_news()
            self.save_to_excel()

    def display_news(self):
        self.text.delete('1.0', tk.END)
        for i, item in enumerate(self.news):
            self.text.insert(tk.END, f'{i + 1}. {item["datetime"]} - {item["title"]} ({item["link"]})\n\n')
            self.text.update_idletasks()

    def save_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Новости'

        for col_num, column_title in enumerate(['Заголовок', 'Дата и время', 'Ссылка'], start=1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, news_elem in enumerate(self.news, start=2):
            ws.cell(row=row_num, column=1).value = news_elem['title']
            ws.cell(row=row_num, column=2).value = news_elem['datetime']
            cell = ws.cell(row=row_num, column=3)
            cell.value = news_elem['link']
            cell.font = Font(color=BLUE, underline='single')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.hyperlink = news_elem['link']

        ws.column_dimensions['A'].width = 105
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 125

        wb.save('rtvi.xlsx')


if __name__ == '__main__':
    root = tk.Tk()
    app = NewsApp(root)
    root.mainloop()