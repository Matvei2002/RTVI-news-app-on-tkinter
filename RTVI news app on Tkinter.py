import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter.ttk import *
from tkinter.scrolledtext import *
from PIL import ImageTk
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLUE
import sys
import os


# Функция для успешного запуска приложения
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# Начало скрипта программы
def get_news(url, num=10):
    # Получаем HTML-код страницы
    response = requests.get(url)
    html_doc = response.text
    soup = BeautifulSoup(html_doc, "html.parser")
    
    # Ищем заголовки, ссылки, даты и времена новостей и выводим указанное количество
    news = []
    for news_elem in soup.find_all('div', class_='arch-block')[:num]:
        # Находим заголовок новости
        title_elem = news_elem.find('h2', class_='arch-title')
        title = title_elem.text.strip()

        # Находим дату и время новости
        datetime_elem = news_elem.find('div', class_='date')
        datetime = datetime_elem.text.strip()

        # Находим ссылку на новость
        link_elem = news_elem.find('a', href=True)
        link = link_elem['href']

        # Добавляем новость в список
        news.append({'title': title, 'datetime': datetime, 'link': link})

    return news


def on_button_click():
    # Получаем адрес новостного сайта из поля ввода
    url = entry.get()

    # Вводим количество новостей
    num = int(input_entry.get())
    
    # Проверяем введенное значение и создаем список новостей 
    if num < 1 or num > 10:
        text.delete('1.0', tk.END)
        text.insert(tk.END,"Ошибка")
        return
    
    else:
        news = get_news(url, num=num)
        text.delete('1.0', tk.END)

        # Выводим список новостей в текстовом поле
        for i, item in enumerate(news):
            text.insert(tk.END, f'{i+1}. {item["datetime"]} - {item["title"]} ({item["link"]})\n\n')
        

    # Записываем результаты в файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Новости'

    for col_num, column_title in enumerate(['Заголовок', 'Дата и время', 'Ссылка'], start=1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, news_elem in enumerate(news, start=2):
        ws.cell(row=row_num, column=1).value = news_elem['title']
        ws.cell(row=row_num, column=2).value = news_elem['datetime']
        cell = ws.cell(row=row_num, column=3)
        cell.value = news_elem['link']
        cell.font = Font(color=BLUE, underline='single')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.hyperlink = news_elem['link']

    # Устанавливаем ширину колонок так, чтобы текст полностью помещался, и сохраняем файл
    ws.column_dimensions['A'].width = 105
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 125

    wb.save('rtvi.xlsx')


# Создаем графический интерфейс и иконку
root = tk.Tk()
root.geometry("800x480")
root.resizable(0,0)
root.config(bg = '#FF9999')
root.title("Новости RTVI")
icon = tk.PhotoImage(file = resource_path('rtvi.png'))
root.iconphoto(False, icon)

# Создаем фрейм под логотип, импортируя его
img = ImageTk.PhotoImage(file = resource_path('rtvi1.png'))
frame = tk.Frame(root)
frame.pack()
label = Label(frame, image = img, background = '#FF9999')
label.pack(side='top')

# Создаем фрейм с заголовком приложения
label_frame = tk.Frame(root)
label_frame.pack()
label = tk.Label(label_frame, text = "Новости", font = "Arial 20 bold", bg = '#FF9999', width = 500)
label.pack(side=  'left')

# Создаем виджет для ввода адреса сайта
entry = tk.Entry(root, font = ('Arial 10'))
entry.pack(side = "top")
entry.insert(0, 'https://rtvi.com/news/')

# Создаем виджет для ввода количества новостей
input_label = tk.Label(root, text = "Количество новостей (от 1 до 10):", font = "Arial 10 bold", bg = '#FF9999')
input_label.pack()
input_entry = tk.Entry(root, width = 10, font = ('Arial 10'))
input_entry.pack(side = "top")

# Создаем кнопку с собственным дизайном для запуска парсинга новостей
style = Style()
style.configure('TButton', font =
               ('Arial 10 bold'),
                foreground = '#DC143C', background = '#FF9999')
button = Button(root, text = "Поиск новостей", style = 'TButton', command = on_button_click)
button.pack(side = "top")

# Создаем фрейм и текстовый виджет для отображения новостей
text_frame = tk.Frame(root)
text_frame.pack(side = "left", fill = "both", expand=True)
text = tk.Text(text_frame, font = ('Times 14'))
text.pack(side = "left", fill = "both", expand=True)

# Создаем скроллбар и привязываем его к текстовому виджету
scrollbar = tk.Scrollbar(text_frame)
scrollbar.pack(side = "right", fill = "y")
scrollbar.config(command = text.yview)
text.config(yscrollcommand = scrollbar.set)


# Запускаем интерфейс приложения
root.mainloop()
